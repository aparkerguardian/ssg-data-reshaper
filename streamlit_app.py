import streamlit as st
import pandas as pd
from github import Github
import io
from io import BytesIO
from openpyxl import load_workbook

# GitHub authentication
access_token = os.getenv('GITHUB_TOKEN')
repository_name = 'aparkerguardian/ssg-data-reshaper'
file_path = 'Mapping.xlsx'

# Connect to the repository
g = Github(access_token)
repo = g.get_repo(repository_name)

def determine_correct_sheet(excel_wb):
    wb = load_workbook(excel_wb)
    sheets = wb.sheetnames
    columns_to_check_for = ['Fund Name', 'Bloomberg Name', 'Category']
    for sheet in sheets:
        df = pd.read_excel(excel_wb, sheet_name=sheet)
        if all(col in df.columns for col in columns_to_check_for):
            return sheet
    return None

def read_github_file():
    try:
        file_content = repo.get_contents(file_path)
        decoded_content = file_content.decoded_content
        sheet = determine_correct_sheet(io.BytesIO(decoded_content))
        return pd.read_excel(io.BytesIO(decoded_content), sheet_name=sheet)
    except Exception as e:
        st.error(f"Error reading file from GitHub: {e}")
        return None

def parse_mapping_file(excel_file):
    sheet = determine_correct_sheet(excel_file)
    if sheet is not None:
        upload_to_github(excel_file, excel_file.name)
    else:
        st.error(f"Uploaded file does not contain the necessary columns: ['Fund Name', 'Bloomberg Name', 'Category']")


def upload_to_github(file, file_name):
    try:
        file_content = file.getvalue()
        # Check if the file already exists in the repository
        try:
            contents = repo.get_contents(file_path)
            repo.update_file(contents.path, "Updating file via Streamlit", file_content, contents.sha)
            st.success(f'File "{file_name}" updated successfully in the GitHub repository.')
        except:
            repo.create_file(file_path, "Uploading file via Streamlit", file_content)
            st.success(f'File "{file_name}" uploaded successfully to the GitHub repository.')
    except Exception as e:
        st.error(f"Error uploading file to GitHub: {e}")



st.markdown('# SSG Data Reformatting Tool')
st.markdown('### Fund Name and Category Mapping')

with st.expander("Click here to preview the current Mapping.xlsx file or replace it with a new one."):
    st.header('Mapping')
    # Allow user to upload a new file to replace the existing one
    uploaded_file = st.file_uploader("Choose a file to replace Mapping.xlsx")
    if uploaded_file is not None:
        if st.button("Upload"):
            parse_mapping_file(uploaded_file)

    # Read and display the current file from GitHub
    df = read_github_file()
    if df is not None:
        st.write("Current Mapping.xlsx from GitHub:")
        st.dataframe(df)

st.markdown('### Upload Eagle PV File')


def pv_reshape(excel_wb):
    
    def parse_mapping_file():
        sheet = determine_correct_sheet('Mapping.xlsx')
        print(sheet)
        lookup = pd.read_excel('Mapping.xlsx', sheet_name=sheet)
        #if lookup doesnt contain the column "Category" then return error
        return lookup
    
    def parse_excel(file_path):
        df = pd.read_excel(file_path)
        df.columns = df.columns.str.rstrip('\n')
        
        #create dummy column to store original values
        df['original'] = df['Category']
        df['original'] = df['original'].ffill().str.strip()
        
        return df
    
    def replace_with_nearest_previous_match(df, lookup):
        last_valid = None
        for index, row in df.iterrows():
            if row['Category'] in lookup['Category'].values:
                last_valid = row['Category']
            else:
                if last_valid is not None:
                    df.at[index, 'Category'] = last_valid
                # Optionally handle case where no previous valid category has been encountered
                # else:
                #     df.at[index, 'Category'] = <default_value_or_logic>
        return df

    def create_output_file():
        
        def replace_with_currency_sum(dfm):
            # Identify the indices of the first occurrences
            index_income_payable = dfm[dfm['original'] == 'INCOME PAYABLE'].index.min() if not dfm[dfm['original'] == 'INCOME PAYABLE'].empty else None
            index_cash = dfm[dfm['original'] == 'CASH'].index.min() if not dfm[dfm['original'] == 'CASH'].empty else None
            index_currency = dfm[dfm['original'] == 'CURRENCY'].index.min() if not dfm[dfm['original'] == 'CURRENCY'].empty else None
            


            # Adjust index_cash if it comes after index_currency
            if index_cash is not None and index_currency is not None and index_cash > index_currency:
                index_cash = None

            # Create a list of the indices to sum from
            indices = [index for index in [index_income_payable, index_cash, index_currency] if index is not None]

            # Sum the 'Market Value Local' values for the selected indices
            total_market_value = dfm.loc[indices, 'Market Value Base'].sum()

            
            #Find the index of the row with the first occurence of "CASH" or "CURRENCY"
            start_index = dfm[
                (dfm['original'] == 'CASH') | 
                (dfm['original'] == 'CURRENCY')
                ].index.min() if not dfm[
                (dfm['original'] == 'CASH') |
                (dfm['original'] == 'CURRENCY')
                ].empty else None
                
            df_filtered = dfm[
                (dfm['original'] == 'INCOME PAYABLE') | 
                (dfm.index >= start_index)
                ] if not dfm[
                (dfm['original'] == 'INCOME PAYABLE') |
                (dfm.index >= start_index)
                ].empty else None
            
            if df_filtered is not None:
                dfm = dfm[~dfm.index.isin(df_filtered.index)].reset_index(drop=True)
            
            
            #new row to be added to the dataframe
            new_row = {'Category': dfm['Category'][0],
                    'Security Number': 'USDCAD Curncy',
                    'Market Value Base': total_market_value,
                    'Effective Date': dfm['Effective Date'][0],
                    'Fund Name': dfm['Fund Name'][0],
                    'Bloomberg Name': dfm['Bloomberg Name'][0],
                    'Security Description 1': 'USDCAD Curncy'
                    }
            
            #append new row
            dfm.loc[len(dfm)] = new_row
            
            return dfm, df_filtered, new_row



        managers = df['Category'].unique()
        #managers = managers[0:1]
        #df = df.dropna(subset=['Security Description 1'])
        df_out = pd.DataFrame(columns=['Effective Date', 'Fund Name', 'Category', 'Security Number', 'Security Description 1', 'Market Value Base', 'Bloomberg Name' ,'Weight'])
        for manager in managers:
            df_manager = df[df['Category'] == manager].copy()
            df_manager = df_manager.merge(lookup, on='Category', how='left')
            df_manager['Effective Date'] = df_manager['As of Date'].dt.strftime('%Y-%m-%d')
            df_manager,  df_filtered, new_row = replace_with_currency_sum(df_manager)
            df_manager = df_manager.dropna(subset=['Security Description 1'])
            df_manager['Weight'] = df_manager['Market Value Base'] / df_manager['Market Value Base'].sum() * 100
            df_manager = df_manager[['Effective Date', 'Fund Name', 'Category', 'Security Number', 'Security Description 1', 'Market Value Base', 'Bloomberg Name' ,'Weight']]
            df_out = pd.concat([df_out, df_manager])

        return df_out

    lookup = parse_mapping_file()
    df = parse_excel(uploaded_pv_file)
    df = replace_with_nearest_previous_match(df, lookup)
    df_out = create_output_file()
    
    
    return df_out



uploaded_pv_file = st.file_uploader("Choose a file to upload")
if uploaded_pv_file is not None:
    if st.button("Reshape PV file"):
        with st.spinner('Creating output file... Please wait.'):
            sheet = determine_correct_sheet('Mapping.xlsx')
            if sheet is not None:
                lookup = pd.read_excel('Mapping.xlsx', sheet_name=sheet)
            else:
                st.error("Please expand the section above and upload a valid Mapping file")
            df_out = pv_reshape(uploaded_pv_file)
            output = BytesIO()
            df_out.to_excel(output, index=False)
            output.seek(0)
        
        st.success('Output file created successfully!')
        st.download_button(
            label="Download output file",
            data=output,
            file_name=f'{uploaded_pv_file.name}_output.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
